# -*- coding: utf-8 -*-
"""
3D3S Excel 导入文件生成器（M3）

按 3D3S"表格化模型导入"模板生成 .xlsx：
    Sheet 节点信息         ：节点编号, X, Y, Z（无表头，mm）
    Sheet 单元信息         ：单元号, 节点1, 节点2, 材料, 截面类型, 截面名称
    Sheet 节点荷载         ：荷载序号, 工况号, 工况名, 类型, Px,Py,Pz,Mx,My,Mz（暂空）
    Sheet 单元信息（含坐标）：单元号, 节点1三坐标, 节点2三坐标, 材料, 截面类型, 截面名称

依赖：openpyxl
"""

from pv_geometry import build_model


# 材料名映射：界面材质 -> 3D3S 材料
MATERIAL_MAP = {
    "Q235B": "Q235",
    "Q355B": "Q355",
    "Q460B": "Q460",
    "Q235": "Q235",
    "Q355": "Q355",
}

# 规格 -> 3D3S 截面类型（依据 3D3S 杆件库"薄壁截面/热轧型钢"页实测名称）
SPEC_3D3S_TYPE = {
    "C型钢":     "冷弯卷边槽钢",
    "Z型钢":     "卷边Z形钢",
    "U型钢":     "U型卷边槽钢",
    "槽钢":      "普通槽钢",        # 热轧型钢页（待实测确认）
    "角钢":      "普通角钢(等肢)",   # 热轧型钢页（模板示例确认）
    "方钢管":    "方形空心型钢",
    "矩形钢管":  "矩形空心型钢",
    "圆钢/圆管": "焊接圆管",        # 待实测确认
}


def _material(params, role):
    sec = params.get("sections", {}).get(role, {})
    mat = sec.get("material", "Q235B")
    return MATERIAL_MAP.get(mat, mat)


def _section(params, role):
    sec = params.get("sections", {}).get(role, {})
    spec = sec.get("spec", "自定义")
    model = sec.get("model", "自定义")
    stype = SPEC_3D3S_TYPE.get(spec, spec)
    # 3D3S 名称用 X 分隔（角钢模板用 x）
    sep = "x" if spec == "角钢" else "X"
    name = model.replace("×", sep).replace("x", sep).replace("X", sep)
    return stype, name


def write_3d3s_excel(params, path):
    """生成 3D3S 表格化模型导入 Excel。返回模型统计信息。"""
    import openpyxl
    from openpyxl import Workbook

    model = build_model(params)
    nodes = model["nodes"]

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "节点信息"
    for idx in sorted(nodes):
        x, y, z = nodes[idx]
        ws1.append([idx, round(x, 3), round(y, 3), round(z, 3)])

    ws2 = wb.create_sheet("单元信息")
    ws4 = wb.create_sheet("单元信息（含坐标）")
    for i, (kind, i1, i2) in enumerate(model["members"], start=1):
        mat = _material(params, kind)
        stype, secname = _section(params, kind)
        x1, y1, z1 = nodes[i1]
        x2, y2, z2 = nodes[i2]
        ws2.append([i, i1, i2, mat, stype, secname])
        ws4.append([i, round(x1, 3), round(y1, 3), round(z1, 3),
                    round(x2, 3), round(y2, 3), round(z2, 3),
                    mat, stype, secname])

    wb.create_sheet("节点荷载")   # 暂空，荷载工况 M4 接入
    wb.save(path)
    return model["info"]


if __name__ == "__main__":
    import sys
    from pv_geometry import default_params
    out = sys.argv[1] if len(sys.argv) > 1 else "光伏支架_3D3S导入.xlsx"
    info = write_3d3s_excel(default_params(), out)
    print("已生成:", out, "| 节点:", info["node_count"], "杆件:", info["member_count"])
